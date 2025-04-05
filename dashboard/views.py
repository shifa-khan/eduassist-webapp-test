from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from rest_framework import permissions, status, views
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import UserRegistrationSerializer, UserSerializer
from django.contrib.auth import authenticate, login
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.exceptions import AuthenticationFailed
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from .models import UploadedFile, ChatMessage
from .forms import FileUploadForm
from django.contrib.auth import logout
from django.http import JsonResponse
import os
from django.core.files.storage import default_storage
from django.views.decorators.csrf import csrf_exempt
import json
from django.conf import settings
from storages.backends.gcloud import GoogleCloudStorage
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import ensure_csrf_cookie
from google.cloud import storage
import datetime
from django.middleware.csrf import get_token
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.views.decorators.http import require_POST
import requests
import mimetypes
from io import BytesIO
import PyPDF2  
import docx    
from openpyxl import load_workbook  
import re

import logging
logger = logging.getLogger(__name__)



# ============================================================
# INDEX VIEW
def index(request):
    """Landing page with Login and Register buttons."""
    return render(request, 'dashboard/index.html') 



# USER REGISTRATION VIEW
def register_view(request):
    """Registers a new user and redirects to login."""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password1')
        
        # Ensure username and email are unique
        if User.objects.filter(username=username).exists():
            return render(request, 'dashboard/register.html', {'error': 'Username already taken'})
        if User.objects.filter(email=email).exists():
            return render(request, 'dashboard/register.html', {'error': 'Email already in use'})

        # DEBUG: Check if User object is created
        try:
            user = User.objects.create_user(username=username, email=email, password=password)
            user.save()
            print(f"User {username} registered successfully!")  # Debugging
            return redirect('login')  # Redirect to login page after registration
        except Exception as e:
            print(f"Registration failed: {e}")  # Print any exception that occurs
            return render(request, 'dashboard/register.html', {'error': 'Registration failed. Try again.'})

    return render(request, 'dashboard/register.html')




# JWT TOKEN GENERATION VIEW
class ObtainTokenPairView(views.APIView):
    """Authenticate user and generate JWT tokens."""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        user = authenticate(username=username, password=password)
        if user is not None:
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            })
        return Response(status=status.HTTP_401_UNAUTHORIZED)



# LOGIN VIEW
def login_view(request):
    """Logs in the user and redirects to their specific dashboard."""
    if request.user.is_authenticated:
        return redirect('user_dashboard', username=request.user.username)  # Redirect already logged-in users

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('user_dashboard', username=user.username)  # Redirects correctly to their own dashboard
        else:
            return render(request, 'dashboard/login.html', {'error': 'Invalid username or password'})

    return render(request, 'dashboard/login.html')


# LOGOUT VIEW
def logout_view(request):
    """Logs out the user and redirects to login."""
    logout(request)
    return redirect('login')



# USER-SPECIFIC DASHBOARD VIEW
@ensure_csrf_cookie
@login_required
def user_dashboard(request, username):
    """User-specific dashboard."""
    if request.user.username != username:
        return redirect('login')  # Prevents unauthorized access

    # For initial GET requests and refresh
    user_files = UploadedFile.objects.filter(user=request.user)
    form = FileUploadForm()

    # Handle form submission
    if request.method == "POST":
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                uploaded_file = form.save(commit=False)
                uploaded_file.user = request.user
                uploaded_file.save()
                print(f"File uploaded to: {uploaded_file.file.url}")
                
                # Redirect to the same page (as a GET request)
                return HttpResponseRedirect(request.path)
            except Exception as e:
                print(f"Error uploading file: {str(e)}")
    
    # Render page (for both initial GET and after redirect)
    response = render(request, 'dashboard/dashboard.html', {"files": user_files, "form": form})
    response.set_cookie("csrftoken", get_token(request))
    return response


# SERVE FILE VIEW (Restricted to File Owner)
def serve_file(request, file_id):
    """Allow only the file owner to access their file."""
    try:
        # First check if user is authenticated via session
        if request.user.is_authenticated:
            user = request.user
        else:
            # Fall back to JWT if session auth is not available
            auth = JWTAuthentication()
            auth_user = auth.authenticate(request)
            if not auth_user:
                return redirect('login')
            user, _ = auth_user

        file = get_object_or_404(UploadedFile, id=file_id)
        
        # Check permissions
        if file.user != user:
            return HttpResponseForbidden("You do not have permission to access this file.")
        
        # Create a public URL that works without signing
        bucket_name = settings.GS_BUCKET_NAME
        file_path = file.file.name
        
        # Create the public URL
        public_url = f"https://storage.googleapis.com/{bucket_name}/{file_path}"
        
        # Redirect to the public URL
        return redirect(public_url)
        
    except AuthenticationFailed:
        return redirect('login')
    except UploadedFile.DoesNotExist:
        return HttpResponseForbidden("File not found.")




# DELETE FILE VIEW
@login_required
def delete_file_view(request, file_id):
    """Delete a file uploaded by the user."""
    try:
        # Authenticate user
        auth = JWTAuthentication()
        user_data = auth.authenticate(request)
        
        if user_data is None:
            # Fallback to session authentication
            user = request.user
        else:
            user, _ = user_data  
        
        # Retrieve the file or return 404 if it doesn't exist
        file = get_object_or_404(UploadedFile, id=file_id, user=user)

        # Delete file
        file.file.delete()  # from GCS
        file.delete()       # from database
        
        return JsonResponse({"message": "File deleted successfully"}, status=200)

    except AuthenticationFailed:
        return JsonResponse({"error": "Invalid credentials"}, status=403)
    except UploadedFile.DoesNotExist:
        return JsonResponse({"error": "File not found"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ChatBot
def get_groq_response(user_message):
    """Get a response from Groq API."""
    try:
        headers = {
            "Authorization": f"Bearer {settings.GROQ_API_KEY}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": "llama3-8b-8192",
            "messages": [
                {"role": "system", "content": "You are an educational assistant helping faculty with syllabus improvement, question paper preparation, grading assessments, etc."},
                {"role": "user", "content": user_message}
            ],
            "temperature": 0.7,
            "max_tokens": 1024,
        }
        
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions", 
            headers=headers, 
            json=payload,
            timeout=30
        )
        
        if response.status_code != 200:
            return f"I'm sorry, there was an error with the AI service. Please try again later."
            
        return response.json()["choices"][0]["message"]["content"]
    except Exception as e:
        return f"I'm sorry, I encountered an error: {str(e)}"

#######################################################
## CHAT BOT FEATURE

# Define allowed file types and size limits
ALLOWED_FILE_TYPES = {
    # Text files
    '.txt': 'text/plain',
    '.csv': 'text/csv',
    '.md': 'text/markdown',
    # Document files
    '.pdf': 'application/pdf',
    '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    # Spreadsheets
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    # Presentations
    '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
    # Add image files
    '.jpg': 'image/jpeg',
    '.jpeg': 'image/jpeg',
    '.png': 'image/png'
}

# Maximum file size (5MB)
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB in bytes

# Maximum tokens for LLM
MAX_TOTAL_TOKENS = 7000
MAX_FILE_TOKENS = 5000

# Helper function for token estimation
def estimate_tokens(text):
    """Rough estimation of token count (4 chars ≈ 1 token)."""
    return len(text) // 4


# File content extraction
def extract_file_content(file_obj):
    """Extract text content from uploaded files."""
    try:
        # Get file extension
        file_name = file_obj.file.name
        file_ext = os.path.splitext(file_name)[1].lower()
        
        # Get file content
        file_content = file_obj.file.read()
        
        # Check file size again as a precaution
        if len(file_content) > MAX_FILE_SIZE:
            return "[File too large for processing]"

        # Image Files
        if file_ext in ['.jpg', '.jpeg', '.png']:
            return f"[This is an image file: {file_name}. The AI cannot directly see the image content, but can discuss it based on your description.]"
        
        # Text extraction based on file type
        if file_ext == '.txt' or file_ext == '.md' or file_ext == '.csv':
            # For text files, decode directly
            return file_content.decode('utf-8', errors='ignore')
            
        elif file_ext == '.pdf':
            try:
                # For PDFs, use PyPDF2
                pdf_reader = PyPDF2.PdfReader(BytesIO(file_content))
                text = ""
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    text += page.extract_text() + "\n"
                return text
            except Exception as e:
                logger.error(f"Error extracting PDF content: {str(e)}")
                return "[Error extracting PDF content]"
                
        elif file_ext == '.docx':
            try:
                # For DOCX, use python-docx
                doc = docx.Document(BytesIO(file_content))
                text = "\n".join([para.text for para in doc.paragraphs if para.text])
                return text
            except Exception as e:
                logger.error(f"Error extracting DOCX content: {str(e)}")
                return "[Error extracting DOCX content]"
                
        elif file_ext == '.xlsx':
            try:
                # For Excel files, use openpyxl
                workbook = load_workbook(BytesIO(file_content), read_only=True)
                text = ""
                # Get the first few sheets
                for sheet_name in workbook.sheetnames[:3]:  # Limit to first 3 sheets
                    sheet = workbook[sheet_name]
                    text += f"Sheet: {sheet_name}\n"
                    # Get first 100 rows and 20 columns max
                    for row in list(sheet.rows)[:100]:
                        row_values = [str(cell.value) if cell.value is not None else "" for cell in row[:20]]
                        text += ", ".join(row_values) + "\n"
                    text += "\n"
                return text
            except Exception as e:
                logger.error(f"Error extracting Excel content: {str(e)}")
                return "[Error extracting Excel content]"
                
        elif file_ext == '.pptx':
            # For PPTX, more complex extraction would be needed
            # Using a placeholder for now
            return "[This is a PowerPoint file. Text extraction requires additional processing.]"
            
        else:
            return f"[File type {file_ext} not supported for content extraction]"
            
    except Exception as e:
        logger.error(f"Error extracting content from file: {str(e)}")
        return "[Error processing file]"
    finally:
        # Reset file pointer position
        if hasattr(file_obj.file, 'seek'):
            file_obj.file.seek(0)

# Handle file uploads for chat
@csrf_exempt
@login_required
def upload_chat_file(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method is allowed'}, status=405)
    
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file provided'}, status=400)
        
        uploaded_file = request.FILES['file']

        ## Debug
        file_ext = os.path.splitext(uploaded_file.name)[1].lower()
        if file_ext in ['.jpg', '.jpeg', '.png']:
            print(f"Processing image file: {uploaded_file.name}")
        
        # Check file size
        if uploaded_file.size > MAX_FILE_SIZE:
            return JsonResponse({
                'error': f'File size exceeds the maximum limit of 5MB'
            }, status=400)
        
        # Check file type
        file_ext = os.path.splitext(uploaded_file.name)[1].lower()
        if file_ext not in ALLOWED_FILE_TYPES:
            return JsonResponse({
                'error': f'File type {file_ext} is not supported. Supported types: {", ".join(ALLOWED_FILE_TYPES.keys())}'
            }, status=400)
        
        # Form processing
        form = FileUploadForm({'category': 'chatbot_files'}, {'file': uploaded_file})
        
        if form.is_valid():
            file_obj = form.save(commit=False)
            file_obj.user = request.user
            file_obj.save()
            
            # Create a public URL manually
            bucket_name = settings.GS_BUCKET_NAME
            file_path = file_obj.file.name
            public_url = f"https://storage.googleapis.com/{bucket_name}/{file_path}"
            
            return JsonResponse({
                'file_id': file_obj.id,
                'file_name': file_obj.file.name,
                'file_url': public_url
            })
        else:
            return JsonResponse({'error': f'Invalid form data: {form.errors}'}, status=400)
    
    except Exception as e:
        logger.error(f"Error uploading chat file: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


# ChatBot API
@csrf_exempt
@require_POST
def chatbot_api(request):
    """API endpoint for chatbot interactions with file support."""
    try:
        data = json.loads(request.body)
        user_message = data.get("message", "")
        file_ids = data.get("file_ids", [])
        
        if not user_message:
            return JsonResponse({"error": "Empty message"}, status=400)
        
        # Get files content for context if file_ids are provided
        file_context = ""
        associated_files = []
        total_tokens = estimate_tokens(user_message)
        
        if file_ids and request.user.is_authenticated:
            for file_id in file_ids:
                try:
                    file_obj = UploadedFile.objects.get(id=file_id, user=request.user)
                    associated_files.append(file_obj)
                    
                    # Extract text content from the file
                    file_content = extract_file_content(file_obj)
                    
                    if file_content:
                        # Estimate tokens for this file content
                        file_tokens = estimate_tokens(file_content)
                        
                        # Check if adding this would exceed our limit
                        if total_tokens + file_tokens > MAX_TOTAL_TOKENS:
                            # Truncate content to fit within limits if possible
                            if file_tokens > 100:  # Only truncate if it's worth it
                                max_chars = (MAX_TOTAL_TOKENS - total_tokens) * 4
                                file_content = file_content[:max_chars] + "...[content truncated due to length]"
                                file_tokens = estimate_tokens(file_content)
                            
                        # Update total tokens and add to context
                        total_tokens += file_tokens
                        file_context += f"\nContent from file '{file_obj.file.name}':\n{file_content}\n"
                        
                        # If we've used up our token budget, stop processing files
                        if total_tokens >= MAX_FILE_TOKENS:
                            file_context += "\n[Additional files were not processed due to length constraints]"
                            break
                            
                except UploadedFile.DoesNotExist:
                    continue
        
        # Prepare enhanced prompt with file context
        enhanced_message = user_message
        if file_context:
            enhanced_message = f"The user asked: {user_message}\n\nI'm also providing content from files they uploaded to help answer their question:{file_context}\n\nBased on this information, please respond to their question."
        
        response_text = get_groq_response(enhanced_message)
        
        # Save to database if user is authenticated
        if request.user.is_authenticated:
            try:
                chat_msg = ChatMessage.objects.create(
                    user=request.user,
                    message=user_message,
                    response=response_text
                )
                # Associate files with this chat message
                if associated_files:
                    chat_msg.files.add(*associated_files)
            except Exception as e:
                logger.error(f"Error saving chat message: {str(e)}")
                # Continue even if saving fails
                pass
        
        return JsonResponse({"response": response_text})
        
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON format"}, status=400)
    except Exception as e:
        logger.error(f"Error in chatbot API: {str(e)}")
        return JsonResponse({"error": str(e)}, status=500)


# Retrive chat history
@login_required
def chat_history(request):
    """Retrieve chat history for the current user."""
    chats = ChatMessage.objects.filter(user=request.user)
    
    # Convert to list of dicts for JSON response
    chat_list = [
        {
            'id': chat.id,
            'message': chat.message,
            'response': chat.response,
            'timestamp': chat.created_at.strftime('%Y-%m-%d %H:%M:%S')
        }
        for chat in chats
    ]
    
    return JsonResponse({'chats': chat_list})